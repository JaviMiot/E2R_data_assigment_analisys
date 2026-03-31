#!/usr/bin/env python3
"""
consolidate_part3_to_csv.py

Consolida todas las adaptaciones E2R de la Práctica III de todos los estudiantes
en un único archivo CSV.
"""

import csv
import json
import logging
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import xlrd


# Configurar directorios
ROOT_DIR = Path(__file__).resolve().parent.parent.parent

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(ROOT_DIR / "logs/part3/part3_consolidation.log", mode="w", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Sheet detection
# ---------------------------------------------------------------------------

def is_part3_sheet(sheet, sheet_name: str) -> bool:
    """Detect whether a sheet contains Part III E2R adaptation data.

    A Part III sheet must contain ALL THREE of these column headers (case-insensitive):
    - "Document Name"
    - "Original Text"
    - "Proposal for E2R Text"

    Args:
        sheet: Sheet object (openpyxl worksheet or xlrd sheet).
        sheet_name: Name of the sheet (used for early-exit heuristic).

    Returns:
        True if the sheet appears to contain Part III data.
    """
    found_doc_name = False
    found_original = False
    found_proposal = False

    try:
        if hasattr(sheet, "iter_rows"):  # openpyxl
            try:
                max_row = min(20, sheet.max_row or 20)
            except Exception:
                max_row = 20

            for i in range(1, max_row + 1):
                rows = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))
                if not rows or not rows[0]:
                    continue
                row_text = " ".join(str(c).lower() for c in rows[0] if c)

                if "document name" in row_text:
                    found_doc_name = True
                if "original text" in row_text:
                    found_original = True
                if "proposal for e2r" in row_text:
                    found_proposal = True
        else:  # xlrd
            max_row = min(20, sheet.nrows)
            for i in range(max_row):
                try:
                    row = sheet.row_values(i)
                    row_text = " ".join(str(c).lower() for c in row if c)

                    if "document name" in row_text:
                        found_doc_name = True
                    if "original text" in row_text:
                        found_original = True
                    if "proposal for e2r" in row_text:
                        found_proposal = True
                except Exception:
                    continue
    except Exception as e:
        logger.debug(f"Error checking sheet '{sheet_name}': {e}")
        return False

    # Require both data-column headers; Document Name header is optional
    # (some students skip the header row and put values directly in row 1)
    return found_original and found_proposal


# ---------------------------------------------------------------------------
# Metadata extraction helpers
# ---------------------------------------------------------------------------

def _find_metadata_and_data_rows_openpyxl(
    sheet,
) -> Tuple[int, Dict[str, str], int]:
    """Locate metadata row and data-header row in an openpyxl sheet.

    The metadata row (index 1-based) contains 'Document Name' in one of its cells.
    The row immediately following it contains the actual metadata values.
    The data-header row contains 'Original Text' and 'Proposal for E2R Text'.

    Returns:
        (meta_value_row, metadata_dict, data_header_row) — 1-based row numbers.
        Returns (0, {}, 0) if not found.
    """
    try:
        max_row = min(30, sheet.max_row or 30)
    except Exception:
        max_row = 30

    meta_header_row = 0
    data_header_row = 0

    for i in range(1, max_row + 1):
        rows = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))
        if not rows or not rows[0]:
            continue
        row = rows[0]
        row_text = " ".join(str(c).lower() for c in row if c)

        if not meta_header_row and "document name" in row_text and "language" in row_text:
            meta_header_row = i
        if "original text" in row_text and "proposal for e2r" in row_text:
            data_header_row = i
            break

    if not data_header_row:
        return 0, {}, 0

    # If no explicit "Document Name" header row was found, assume row 1 holds
    # the metadata values directly (col order: Document Name, Language, Tool)
    if not meta_header_row:
        value_row_idx = 1
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if first_row and first_row[0]:
            vals = first_row[0]
            fallback_meta = {
                "Document Name": str(vals[0]).strip() if len(vals) > 0 and vals[0] else "",
                "Language": str(vals[1]).strip() if len(vals) > 1 and vals[1] else "",
                "Tool used for the adaptation": str(vals[2]).strip() if len(vals) > 2 and vals[2] else "",
            }
            # Sanitise "None" strings
            for k in fallback_meta:
                if fallback_meta[k] == "None":
                    fallback_meta[k] = ""
            return value_row_idx, fallback_meta, data_header_row
        return 0, {}, 0

    # Values row is the one right after the header row
    value_row_idx = meta_header_row + 1
    value_rows = list(
        sheet.iter_rows(min_row=value_row_idx, max_row=value_row_idx, values_only=True)
    )
    header_rows = list(
        sheet.iter_rows(min_row=meta_header_row, max_row=meta_header_row, values_only=True)
    )

    metadata: Dict[str, str] = {}
    if header_rows and header_rows[0] and value_rows and value_rows[0]:
        headers = header_rows[0]
        values = value_rows[0]
        for col_idx, hdr in enumerate(headers):
            if hdr and col_idx < len(values):
                key = str(hdr).strip()
                val = str(values[col_idx]).strip() if values[col_idx] is not None else ""
                if val == "None":
                    val = ""
                metadata[key] = val

    return value_row_idx, metadata, data_header_row


def _find_metadata_and_data_rows_xlrd(
    sheet,
) -> Tuple[int, Dict[str, str], int]:
    """Locate metadata row and data-header row in an xlrd sheet (0-indexed).

    Returns:
        (meta_value_row, metadata_dict, data_header_row) — 0-based row numbers.
        Returns (0, {}, 0) if not found.
    """
    max_row = min(30, sheet.nrows)
    meta_header_row = -1
    data_header_row = -1

    for i in range(max_row):
        try:
            row = sheet.row_values(i)
            row_text = " ".join(str(c).lower() for c in row if c)

            if meta_header_row == -1 and "document name" in row_text and "language" in row_text:
                meta_header_row = i
            if "original text" in row_text and "proposal for e2r" in row_text:
                data_header_row = i
                break
        except Exception:
            continue

    if data_header_row == -1:
        return 0, {}, 0

    # If no explicit "Document Name" header row, assume row 0 holds values directly
    if meta_header_row == -1:
        if sheet.nrows > 0:
            vals = sheet.row_values(0)
            fallback_meta = {
                "Document Name": str(vals[0]).strip() if len(vals) > 0 and vals[0] else "",
                "Language": str(vals[1]).strip() if len(vals) > 1 and vals[1] else "",
                "Tool used for the adaptation": str(vals[2]).strip() if len(vals) > 2 and vals[2] else "",
            }
            return 0, fallback_meta, data_header_row
        return 0, {}, 0

    value_row_idx = meta_header_row + 1
    headers = sheet.row_values(meta_header_row)
    values = sheet.row_values(value_row_idx) if value_row_idx < sheet.nrows else []

    metadata: Dict[str, str] = {}
    for col_idx, hdr in enumerate(headers):
        if hdr and col_idx < len(values):
            key = str(hdr).strip()
            val = str(values[col_idx]).strip() if values[col_idx] else ""
            metadata[key] = val

    return value_row_idx, metadata, data_header_row


# ---------------------------------------------------------------------------
# Multi-tool column detection
# ---------------------------------------------------------------------------

def _detect_multi_tool_groups_openpyxl(sheet, data_header_row: int) -> List[Dict]:
    """Detect column groups for multi-tool side-by-side layouts (openpyxl).

    When a sheet contains two document analyses side by side, there is a None
    separator column between groups.  Each group has the structure:
      [Original Text] [Proposal for E2R Text] [Comments?]

    Args:
        sheet: openpyxl worksheet.
        data_header_row: 1-based row number of the data header row.

    Returns:
        List of dicts with keys 'orig_col', 'proposal_col', 'comments_col' (1-based).
        Returns a single group with the first matching columns if no multi-layout found.
    """
    header_cells = list(
        sheet.iter_rows(
            min_row=data_header_row, max_row=data_header_row, values_only=True
        )
    )[0]

    groups = []
    current_orig: Optional[int] = None
    current_proposal: Optional[int] = None
    current_comments: Optional[int] = None

    for col_idx, cell in enumerate(header_cells, start=1):
        if cell is None:
            # Separator — if we have a partial group, save it
            if current_orig is not None and current_proposal is not None:
                groups.append(
                    {
                        "orig_col": current_orig,
                        "proposal_col": current_proposal,
                        "comments_col": current_comments,
                    }
                )
                current_orig = None
                current_proposal = None
                current_comments = None
            continue

        cell_text = str(cell).lower().strip()
        if "original text" in cell_text:
            current_orig = col_idx
        elif "proposal for e2r" in cell_text:
            current_proposal = col_idx
        elif "comment" in cell_text:
            current_comments = col_idx

    # Save last group
    if current_orig is not None and current_proposal is not None:
        groups.append(
            {
                "orig_col": current_orig,
                "proposal_col": current_proposal,
                "comments_col": current_comments,
            }
        )

    return groups if groups else [{"orig_col": 1, "proposal_col": 2, "comments_col": 3}]


def _detect_multi_tool_groups_xlrd(sheet, data_header_row: int) -> List[Dict]:
    """Detect column groups for multi-tool side-by-side layouts (xlrd).

    Args:
        sheet: xlrd sheet object.
        data_header_row: 0-based row index of the data header row.

    Returns:
        List of dicts with keys 'orig_col', 'proposal_col', 'comments_col' (0-based).
    """
    header_cells = sheet.row_values(data_header_row)

    groups = []
    current_orig: Optional[int] = None
    current_proposal: Optional[int] = None
    current_comments: Optional[int] = None

    for col_idx, cell in enumerate(header_cells):
        if not cell:
            if current_orig is not None and current_proposal is not None:
                groups.append(
                    {
                        "orig_col": current_orig,
                        "proposal_col": current_proposal,
                        "comments_col": current_comments,
                    }
                )
                current_orig = None
                current_proposal = None
                current_comments = None
            continue

        cell_text = str(cell).lower().strip()
        if "original text" in cell_text:
            current_orig = col_idx
        elif "proposal for e2r" in cell_text:
            current_proposal = col_idx
        elif "comment" in cell_text:
            current_comments = col_idx

    if current_orig is not None and current_proposal is not None:
        groups.append(
            {
                "orig_col": current_orig,
                "proposal_col": current_proposal,
                "comments_col": current_comments,
            }
        )

    return groups if groups else [{"orig_col": 0, "proposal_col": 1, "comments_col": 2}]


# ---------------------------------------------------------------------------
# Metadata per group (multi-tool support)
# ---------------------------------------------------------------------------

def _extract_group_metadata_openpyxl(
    sheet,
    meta_value_row: int,
    meta_header_row: int,
    orig_col: int,
) -> Dict[str, str]:
    """Extract Document Name / Language / Tool for one column group (openpyxl).

    For multi-tool layouts the metadata rows span multiple column groups.
    We read the meta header and value rows starting from the column of 'Original Text'.

    Args:
        sheet: openpyxl worksheet.
        meta_value_row: 1-based row with metadata values.
        meta_header_row: 1-based row with header labels.
        orig_col: 1-based starting column of this group's 'Original Text'.

    Returns:
        Dict with keys 'Document Name', 'Language', 'Tool used for the adaptation'.
    """
    # For single-group sheets, Document Name is typically at col 1.
    # For multi-group sheets, the group's metadata starts at orig_col - 2 (or col 1).
    # Strategy: find the 'Document Name' cell in the header row that is closest to
    # but not past orig_col.
    try:
        max_col = sheet.max_column or 10
    except Exception:
        max_col = 10

    header_row_vals = list(
        sheet.iter_rows(min_row=meta_header_row, max_row=meta_header_row, values_only=True)
    )[0]
    value_row_vals = list(
        sheet.iter_rows(min_row=meta_value_row, max_row=meta_value_row, values_only=True)
    )[0]

    # Find the Document Name column nearest to orig_col
    doc_name_col = None
    for col_idx, cell in enumerate(header_row_vals, start=1):
        if cell and "document name" in str(cell).lower():
            if col_idx <= orig_col:
                doc_name_col = col_idx

    if doc_name_col is None:
        doc_name_col = 1

    meta: Dict[str, str] = {}
    for offset, key in enumerate(
        ["Document Name", "Language", "Tool used for the adaptation"]
    ):
        col = doc_name_col + offset - 1  # 0-based index
        val = value_row_vals[col] if col < len(value_row_vals) else None
        meta[key] = str(val).strip() if val is not None and str(val) != "None" else ""

    # Some students list multiple tools across multiple rows — collect them
    # by scanning additional value rows until we hit the data header
    return meta


def _extract_group_metadata_xlrd(
    sheet,
    meta_value_row: int,
    meta_header_row: int,
    orig_col: int,
) -> Dict[str, str]:
    """Extract Document Name / Language / Tool for one column group (xlrd).

    Args:
        sheet: xlrd sheet.
        meta_value_row: 0-based row with metadata values.
        meta_header_row: 0-based row with header labels.
        orig_col: 0-based starting column of this group's 'Original Text'.

    Returns:
        Dict with keys 'Document Name', 'Language', 'Tool used for the adaptation'.
    """
    header_row_vals = sheet.row_values(meta_header_row)
    value_row_vals = sheet.row_values(meta_value_row) if meta_value_row < sheet.nrows else []

    doc_name_col = None
    for col_idx, cell in enumerate(header_row_vals):
        if cell and "document name" in str(cell).lower():
            if col_idx <= orig_col:
                doc_name_col = col_idx

    if doc_name_col is None:
        doc_name_col = 0

    meta: Dict[str, str] = {}
    for offset, key in enumerate(
        ["Document Name", "Language", "Tool used for the adaptation"]
    ):
        col = doc_name_col + offset
        val = value_row_vals[col] if col < len(value_row_vals) else None
        meta[key] = str(val).strip() if val else ""

    return meta


# ---------------------------------------------------------------------------
# Tool value consolidation (some students list tools on multiple rows)
# ---------------------------------------------------------------------------

def _collect_tool_rows_openpyxl(
    sheet, meta_value_row: int, data_header_row: int, tool_col_idx: int
) -> str:
    """Collect tool values from rows between meta_value_row and data_header_row.

    Some students list multiple tools on consecutive rows (e.g., row 2: 'ChatGPT',
    row 3: 'Gemini').  This joins them with ', '.

    Args:
        sheet: openpyxl worksheet.
        meta_value_row: 1-based row of first metadata values.
        data_header_row: 1-based row of data header.
        tool_col_idx: 0-based column index for tool value.

    Returns:
        Joined tool string.
    """
    tools = []
    for r in range(meta_value_row, data_header_row):
        row_vals = list(
            sheet.iter_rows(min_row=r, max_row=r, values_only=True)
        )[0]
        val = row_vals[tool_col_idx] if tool_col_idx < len(row_vals) else None
        if val and str(val).strip() and str(val).strip() != "None":
            tools.append(str(val).strip())
    return ", ".join(tools) if tools else ""


def _collect_tool_rows_xlrd(
    sheet, meta_value_row: int, data_header_row: int, tool_col_idx: int
) -> str:
    """Collect tool values from rows between meta_value_row and data_header_row (xlrd).

    Args:
        sheet: xlrd sheet.
        meta_value_row: 0-based row of first metadata values.
        data_header_row: 0-based row of data header.
        tool_col_idx: 0-based column index for tool value.

    Returns:
        Joined tool string.
    """
    tools = []
    for r in range(meta_value_row, data_header_row):
        if r >= sheet.nrows:
            break
        row_vals = sheet.row_values(r)
        val = row_vals[tool_col_idx] if tool_col_idx < len(row_vals) else None
        if val and str(val).strip():
            tools.append(str(val).strip())
    return ", ".join(tools) if tools else ""


# ---------------------------------------------------------------------------
# Data row extraction
# ---------------------------------------------------------------------------

def _extract_rows_from_group_openpyxl(
    sheet,
    data_header_row: int,
    group_meta: Dict[str, str],
    col_group: Dict,
    group_num: int,
) -> List[Dict[str, str]]:
    """Extract data rows for one column group from an openpyxl sheet.

    Args:
        sheet: openpyxl worksheet.
        data_header_row: 1-based row of the data header.
        group_meta: Dict with Document Name, Language, Tool.
        col_group: Dict with orig_col, proposal_col, comments_col (1-based).
        group_num: Student group number.

    Returns:
        List of CSV row dicts.
    """
    orig_parts = []
    proposal_parts = []
    detailed_comments = []

    orig_col = col_group["orig_col"]
    proposal_col = col_group["proposal_col"]
    comments_col = col_group.get("comments_col")

    try:
        max_row = sheet.max_row or data_header_row + 1
    except Exception:
        max_row = data_header_row + 1000  # Aumentar margen

    for r in range(data_header_row + 1, max_row + 1):
        row_vals = list(sheet.iter_rows(min_row=r, max_row=r, values_only=True))
        if not row_vals or not row_vals[0]:
            continue
        row = row_vals[0]

        orig = row[orig_col - 1] if orig_col - 1 < len(row) else None
        proposal = row[proposal_col - 1] if proposal_col - 1 < len(row) else None
        comment = (
            row[comments_col - 1]
            if comments_col is not None and comments_col - 1 < len(row)
            else None
        )

        orig_str = str(orig).strip() if orig is not None else ""
        proposal_str = str(proposal).strip() if proposal is not None else ""
        comment_str = str(comment).strip() if comment is not None and str(comment) != "None" else ""

        if not orig_str and not proposal_str and not comment_str:
            continue

        if orig_str:
            orig_parts.append(orig_str)
        if proposal_str:
            proposal_parts.append(proposal_str)
        if comment_str:
            detailed_comments.append({
                "orig": orig_str,
                "prop": proposal_str,
                "comment": comment_str
            })

    if not orig_parts and not proposal_parts and not detailed_comments:
        return []

    full_original = " ".join(orig_parts)
    full_proposal = " ".join(proposal_parts)
    comments_json = json.dumps(detailed_comments, ensure_ascii=False) if detailed_comments else ""

    return [
        {
            "Group": str(group_num),
            "Document Name": group_meta.get("Document Name", ""),
            "Language": group_meta.get("Language", ""),
            "Tool used for the adaptation": group_meta.get(
                "Tool used for the adaptation", ""
            ),
            "Original Text": full_original,
            "Proposal for E2R Text": full_proposal,
            "Comments": comments_json,
        }
    ]


def _extract_rows_from_group_xlrd(
    sheet,
    data_header_row: int,
    group_meta: Dict[str, str],
    col_group: Dict,
    group_num: int,
) -> List[Dict[str, str]]:
    """Extract data rows for one column group from an xlrd sheet.

    Args:
        sheet: xlrd sheet.
        data_header_row: 0-based row of the data header.
        group_meta: Dict with Document Name, Language, Tool.
        col_group: Dict with orig_col, proposal_col, comments_col (0-based).
        group_num: Student group number.

    Returns:
        List of CSV row dicts.
    """
    orig_parts = []
    proposal_parts = []
    detailed_comments = []

    orig_col = col_group["orig_col"]
    proposal_col = col_group["proposal_col"]
    comments_col = col_group.get("comments_col")

    for r in range(data_header_row + 1, sheet.nrows):
        try:
            row = sheet.row_values(r)
        except Exception:
            continue

        orig = row[orig_col] if orig_col < len(row) else None
        proposal = row[proposal_col] if proposal_col < len(row) else None
        comment = (
            row[comments_col]
            if comments_col is not None and comments_col < len(row)
            else None
        )

        orig_str = str(orig).strip() if orig else ""
        proposal_str = str(proposal).strip() if proposal else ""
        comment_str = str(comment).strip() if comment else ""

        if not orig_str and not proposal_str and not comment_str:
            continue

        if orig_str:
            orig_parts.append(orig_str)
        if proposal_str:
            proposal_parts.append(proposal_str)
        if comment_str:
            detailed_comments.append({
                "orig": orig_str,
                "prop": proposal_str,
                "comment": comment_str
            })

    if not orig_parts and not proposal_parts and not detailed_comments:
        return []

    full_original = " ".join(orig_parts)
    full_proposal = " ".join(proposal_parts)
    comments_json = json.dumps(detailed_comments, ensure_ascii=False) if detailed_comments else ""

    return [
        {
            "Group": str(group_num),
            "Document Name": group_meta.get("Document Name", ""),
            "Language": group_meta.get("Language", ""),
            "Tool used for the adaptation": group_meta.get(
                "Tool used for the adaptation", ""
            ),
            "Original Text": full_original,
            "Proposal for E2R Text": full_proposal,
            "Comments": comments_json,
        }
    ]


# ---------------------------------------------------------------------------
# Sheet processing (openpyxl and xlrd variants)
# ---------------------------------------------------------------------------

def _process_sheet_openpyxl(sheet, sheet_name: str, group_num: int) -> List[Dict[str, str]]:
    """Process a single openpyxl Part III sheet.

    Args:
        sheet: openpyxl worksheet confirmed to be a Part III sheet.
        sheet_name: Name of the sheet (for logging).
        group_num: Student group number.

    Returns:
        List of CSV row dicts.
    """
    meta_value_row, metadata, data_header_row = _find_metadata_and_data_rows_openpyxl(sheet)
    if not data_header_row:
        logger.warning(f"    Could not locate data header row in sheet '{sheet_name}'")
        return []

    meta_header_row = meta_value_row - 1

    col_groups = _detect_multi_tool_groups_openpyxl(sheet, data_header_row)
    logger.debug(f"    Sheet '{sheet_name}': {len(col_groups)} column group(s)")

    all_rows = []
    for col_group in col_groups:
        orig_col = col_group["orig_col"]

        # Determine 0-based tool column for multi-row tool collection
        # Tool column = orig_col - 1 (0-based) + 2 offset back from orig
        # Actually: Document Name is 2 cols before Original Text in standard layout
        # doc_name_col (1-based) = orig_col - 2 for standard, = 1 for first group
        # We find it from the metadata extraction
        group_meta = _extract_group_metadata_openpyxl(
            sheet, meta_value_row, meta_header_row, orig_col
        )

        # Collect multi-row tool values (some students list tools on multiple rows)
        # Find the tool column index (0-based)
        header_row_vals = list(
            sheet.iter_rows(
                min_row=meta_header_row, max_row=meta_header_row, values_only=True
            )
        )[0]
        tool_col_0based = None
        doc_name_col_1based = None
        for ci, cell in enumerate(header_row_vals, start=1):
            if cell and "document name" in str(cell).lower() and ci <= orig_col:
                doc_name_col_1based = ci
        if doc_name_col_1based is not None:
            tool_col_0based = doc_name_col_1based + 1  # 0-based = col 3 in group (index 2)

        if tool_col_0based is not None:
            tool_str = _collect_tool_rows_openpyxl(
                sheet, meta_value_row, data_header_row, tool_col_0based
            )
            if tool_str:
                group_meta["Tool used for the adaptation"] = tool_str

        rows = _extract_rows_from_group_openpyxl(
            sheet, data_header_row, group_meta, col_group, group_num
        )
        logger.debug(f"      Group col {orig_col}: {len(rows)} rows, doc='{group_meta.get('Document Name','')[:40]}'")
        all_rows.extend(rows)

    return all_rows


def _process_sheet_xlrd(sheet, sheet_name: str, group_num: int) -> List[Dict[str, str]]:
    """Process a single xlrd Part III sheet.

    Args:
        sheet: xlrd sheet confirmed to be a Part III sheet.
        sheet_name: Name of the sheet (for logging).
        group_num: Student group number.

    Returns:
        List of CSV row dicts.
    """
    meta_value_row, metadata, data_header_row = _find_metadata_and_data_rows_xlrd(sheet)
    if not data_header_row:
        logger.warning(f"    Could not locate data header row in sheet '{sheet_name}'")
        return []

    meta_header_row = meta_value_row - 1

    col_groups = _detect_multi_tool_groups_xlrd(sheet, data_header_row)
    logger.debug(f"    Sheet '{sheet_name}': {len(col_groups)} column group(s)")

    all_rows = []
    for col_group in col_groups:
        orig_col = col_group["orig_col"]

        group_meta = _extract_group_metadata_xlrd(
            sheet, meta_value_row, meta_header_row, orig_col
        )

        header_row_vals = sheet.row_values(meta_header_row)
        tool_col_0based = None
        doc_name_col_0based = None
        for ci, cell in enumerate(header_row_vals):
            if cell and "document name" in str(cell).lower() and ci <= orig_col:
                doc_name_col_0based = ci
        if doc_name_col_0based is not None:
            tool_col_0based = doc_name_col_0based + 2  # 0-based col for Tool

        if tool_col_0based is not None:
            tool_str = _collect_tool_rows_xlrd(
                sheet, meta_value_row, data_header_row, tool_col_0based
            )
            if tool_str:
                group_meta["Tool used for the adaptation"] = tool_str

        rows = _extract_rows_from_group_xlrd(
            sheet, data_header_row, group_meta, col_group, group_num
        )
        all_rows.extend(rows)

    return all_rows


# ---------------------------------------------------------------------------
# Excel file processing
# ---------------------------------------------------------------------------

def process_excel_file_part3(file_path: Path, group_num: int) -> List[Dict[str, str]]:
    """Process one Excel file and return all Part III rows found.

    Tries openpyxl first; falls back to xlrd for genuine .xls files.
    Handles .XLS files that are actually xlsx by using a temp copy.

    Args:
        file_path: Path to the Excel file.
        group_num: Student group number.

    Returns:
        List of CSV row dicts extracted from any Part III sheets found.
    """
    all_rows: List[Dict[str, str]] = []
    suffix = file_path.suffix.lower()

    # --- Try openpyxl first ---
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if is_part3_sheet(ws, sheet_name):
                logger.info(f"    → Part III sheet found: '{sheet_name}'")
                rows = _process_sheet_openpyxl(ws, sheet_name, group_num)
                logger.info(f"      {len(rows)} rows extracted")
                all_rows.extend(rows)
        wb.close()
        return all_rows
    except Exception as openpyxl_err:
        logger.debug(f"    openpyxl failed for {file_path.name}: {openpyxl_err}")

    # --- Try .XLS disguised as xlsx (copy to .xlsx temp file) ---
    if suffix in (".xls", ".XLS".lower()):
        tmp_path: Optional[Path] = None
        try:
            tmp_path = Path(tempfile.mktemp(suffix=".xlsx"))
            shutil.copy(file_path, tmp_path)
            wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=False)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if is_part3_sheet(ws, sheet_name):
                    logger.info(f"    → Part III sheet found (XLS→xlsx): '{sheet_name}'")
                    rows = _process_sheet_openpyxl(ws, sheet_name, group_num)
                    logger.info(f"      {len(rows)} rows extracted")
                    all_rows.extend(rows)
            wb.close()
            return all_rows
        except Exception as xlsx_disguised_err:
            logger.debug(f"    XLS→xlsx attempt failed for {file_path.name}: {xlsx_disguised_err}")
        finally:
            if tmp_path and tmp_path.exists():
                tmp_path.unlink()

    # --- Try xlrd for genuine old .xls ---
    try:
        wb = xlrd.open_workbook(str(file_path))
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            if is_part3_sheet(ws, sheet_name):
                logger.info(f"    → Part III sheet found (xlrd): '{sheet_name}'")
                rows = _process_sheet_xlrd(ws, sheet_name, group_num)
                logger.info(f"      {len(rows)} rows extracted")
                all_rows.extend(rows)
        return all_rows
    except Exception as xlrd_err:
        logger.warning(f"    All read attempts failed for {file_path.name}: {xlrd_err}")

    return all_rows


# ---------------------------------------------------------------------------
# Main scan
# ---------------------------------------------------------------------------

def find_all_part3_data() -> Tuple[List[Dict[str, str]], Dict[str, int], List[str]]:
    """Scan all student directories and collect Part III data.

    Returns:
        (all_rows, group_mapping, groups_without_data)
    """
    base_dir = Path("/home/javi/practices/practices")
    student_dirs = sorted([d for d in base_dir.glob("*_assignsubmission_file")])

    # Same group numbering as consolidate_part2_to_csv.py
    group_mapping = {d.name: idx + 1 for idx, d in enumerate(student_dirs)}

    all_rows: List[Dict[str, str]] = []
    groups_without_data: List[str] = []

    logger.info(f"Found {len(student_dirs)} student directories")
    logger.info("=" * 80)

    for student_dir in student_dirs:
        student_name = student_dir.name.split("_")[0]
        group_num = group_mapping[student_dir.name]

        logger.info(f"\nGroup {group_num}: {student_name}")

        # Collect all Excel files (case-insensitive)
        xlsx_files = list(student_dir.glob("*.xlsx")) + list(student_dir.glob("*.XLSX"))
        xls_files = list(student_dir.glob("*.xls")) + list(student_dir.glob("*.XLS"))
        all_excel = xlsx_files + xls_files

        if not all_excel:
            logger.warning(f"  No Excel files found")
            groups_without_data.append(f"{group_num}. {student_name}")
            continue

        student_rows: List[Dict[str, str]] = []
        for excel_file in all_excel:
            # Skip macOS metadata files
            if excel_file.name.startswith("._") or excel_file.name.startswith(".~"):
                continue
            logger.info(f"  File: {excel_file.name}")
            rows = process_excel_file_part3(excel_file, group_num)
            student_rows.extend(rows)

        if student_rows:
            all_rows.extend(student_rows)
            logger.info(f"  ✓ Total rows for this group: {len(student_rows)}")
        else:
            logger.warning(f"  ✗ No Part III data found")
            groups_without_data.append(f"{group_num}. {student_name}")

    return all_rows, group_mapping, groups_without_data


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_csv_part3(all_rows: List[Dict[str, str]], output_path: Path) -> None:
    """Write the consolidated Part III CSV.

    Args:
        all_rows: List of row dicts.
        output_path: Destination CSV path.
    """
    if not all_rows:
        logger.error("No data to write!")
        return

    headers = [
        "Group",
        "Document Name",
        "Language",
        "Tool used for the adaptation",
        "Original Text",
        "Proposal for E2R Text",
        "Comments",
    ]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_rows)

    logger.info(f"\n✅ CSV written to: {output_path}")
    logger.info(f"   Total rows: {len(all_rows)}")


def write_report_part3(
    all_rows: List[Dict[str, str]],
    group_mapping: Dict[str, int],
    groups_without_data: List[str],
    report_path: Path,
) -> None:
    """Generate a consolidation report.

    Args:
        all_rows: All extracted rows.
        group_mapping: Dir-name → group number mapping.
        groups_without_data: List of groups with no Part III data.
        report_path: Destination report file path.
    """
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("=" * 80 + "\n")
        f.write("REPORTE DE CONSOLIDACIÓN - PRÁCTICA III\n")
        f.write("=" * 80 + "\n\n")

        f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        f.write("ESTADÍSTICAS:\n")
        f.write(f"  • Total de grupos: {len(group_mapping)}\n")
        f.write(
            f"  • Grupos con datos: {len(group_mapping) - len(groups_without_data)}\n"
        )
        f.write(f"  • Grupos sin datos: {len(groups_without_data)}\n")
        f.write(f"  • Total de filas CSV: {len(all_rows)}\n\n")

        if groups_without_data:
            f.write("⚠️  GRUPOS SIN PRÁCTICA III:\n")
            for group in groups_without_data:
                f.write(f"  {group}\n")
            f.write("\n")

        f.write("\nDISTRIBUCIÓN DE FILAS POR GRUPO:\n")
        group_counts: Dict[str, int] = {}
        for row in all_rows:
            g = row["Group"]
            group_counts[g] = group_counts.get(g, 0) + 1

        for group_num in sorted(group_counts.keys(), key=int):
            count = group_counts[group_num]
            f.write(f"  Grupo {group_num}: {count} filas\n")

        f.write("\n" + "=" * 80 + "\n")

    logger.info(f"📄 Report written to: {report_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    """Main entry point."""
    logger.info("Starting Part III consolidation...")
    logger.info("=" * 80)

    all_rows, group_mapping, groups_without_data = find_all_part3_data()

    logger.info("\n" + "=" * 80)
    logger.info("CONSOLIDATION COMPLETE")
    logger.info("=" * 80)

    if not all_rows:
        logger.error("❌ No data found to consolidate!")
        return

    csv_path = ROOT_DIR / "data/part3/part3_consolidated.csv"
    write_csv_part3(all_rows, csv_path)

    report_path = ROOT_DIR / "logs/part3/part3_consolidation_report.txt"
    write_report_part3(all_rows, group_mapping, groups_without_data, report_path)

    logger.info("\n✅ All done!")
    logger.info(f"   CSV:    {csv_path}")
    logger.info(f"   Report: {report_path}")
    logger.info(f"   Log:    logs/part3/part3_consolidation.log")


if __name__ == "__main__":
    main()
