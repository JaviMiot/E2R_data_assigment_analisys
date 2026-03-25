#!/usr/bin/env python3
"""
consolidate_part2_to_csv.py

Consolida todas las hojas de Práctica II de todos los estudiantes
en un único archivo CSV.
"""

import openpyxl
import xlrd
import csv
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime


# Configurar directorios
ROOT_DIR = Path(__file__).resolve().parent.parent

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(ROOT_DIR / "logs/part2_consolidation.log", mode="w", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)

logger = logging.getLogger(__name__)


def is_part2_sheet(sheet, sheet_name: str) -> bool:
    """
    Detecta si una hoja contiene datos Part II.

    Criterios:
    1. NO debe ser una hoja "Values" (son resúmenes)
    2. Debe contener al menos 2 de 3 indicadores:
       - "Elements to be identified"
       - "Value" como header
       - "E2R Guideline" o "E2R Analysis"
    """
    # Excluir hojas de resumen
    if "value" in sheet_name.lower() and sheet_name.lower().startswith("value"):
        return False

    found_elements = False
    found_value = False
    found_e2r = False

    try:
        # Manejar tanto openpyxl como xlrd
        if hasattr(sheet, "iter_rows"):  # openpyxl
            # Para openpyxl, obtener max_row de forma segura
            try:
                max_row = min(20, sheet.max_row or 20)
            except:
                max_row = 20

            for i in range(1, max_row + 1):
                row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))
                if row and row[0]:
                    row_text = " ".join(
                        [str(cell).lower() if cell else "" for cell in row[0]]
                    )

                    if "elements to be identified" in row_text:
                        found_elements = True
                    if "value" in row_text and i < 15:
                        found_value = True
                    if "e2r guideline" in row_text or "e2r analysis" in row_text:
                        found_e2r = True
        else:  # xlrd
            max_row = min(20, sheet.nrows)
            for i in range(max_row):
                try:
                    row = sheet.row_values(i)
                    row_text = " ".join(
                        [str(cell).lower() if cell else "" for cell in row]
                    )

                    if "elements to be identified" in row_text:
                        found_elements = True
                    if "value" in row_text and i < 15:
                        found_value = True
                    if "e2r guideline" in row_text or "e2r analysis" in row_text:
                        found_e2r = True
                except:
                    continue
    except Exception as e:
        logger.debug(f"Error checking sheet {sheet_name}: {e}")
        return False

    score = sum([found_elements, found_value, found_e2r])
    return score >= 2


def extract_metadata_openpyxl(sheet) -> Dict[str, str]:
    """Extrae metadatos de las primeras filas (openpyxl)."""
    metadata = {"web": "", "evaluators": "", "tool": "", "comment": "", "time": ""}

    # Buscar en primeras 10 filas
    for row in sheet.iter_rows(max_row=10, values_only=True):
        if not row or not row[0]:
            continue

        key = str(row[0]).strip().lower()
        value = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if "web site" in key:
            metadata["web"] = value
        elif "evaluator" in key:
            metadata["evaluators"] = value
        elif "tool used" in key:
            metadata["tool"] = value
        elif "comment" in key:
            metadata["comment"] = value
        elif "time spent" in key:
            metadata["time"] = value

    return metadata


def extract_metadata_before_table_openpyxl(sheet, header_row: int) -> Dict[str, str]:
    """
    Extrae metadatos específicos de las filas ANTES de una tabla específica.

    Args:
        sheet: La hoja de Excel
        header_row: Fila donde está el header de la tabla

    Returns:
        Diccionario con metadata incluyendo tool_indicator si se encuentra
    """
    metadata = {
        "web": "",
        "evaluators": "",
        "tool": "",
        "comment": "",
        "time": "",
        "tool_indicator": "",
    }

    # Buscar metadata en las 20 filas anteriores al header
    search_start = max(1, header_row - 20)

    for i in range(search_start, header_row):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if not row or not row[0]:
            continue

        key = str(row[0]).strip().lower()
        value = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if "web site" in key:
            metadata["web"] = value
        elif "evaluator" in key:
            metadata["evaluators"] = value
        elif "tool used" in key:
            metadata["tool"] = value
        elif "comment" in key:
            metadata["comment"] = value
        elif "time spent" in key:
            metadata["time"] = value

    # Buscar indicador de herramienta específico en la fila INMEDIATAMENTE anterior al header
    if header_row > 1:
        prev_row = list(
            sheet.iter_rows(
                min_row=header_row - 1, max_row=header_row - 1, values_only=True
            )
        )[0]
        if len(prev_row) > 1 and prev_row[1]:
            tool_indicator = str(prev_row[1]).strip()
            # Verificar si contiene nombre de herramienta
            tool_lower = tool_indicator.lower()
            if any(
                keyword in tool_lower
                for keyword in ["gemini", "chatgpt", "chat gpt", "claude", "gpt"]
            ):
                metadata["tool_indicator"] = tool_indicator
                logger.debug(f"      Tool indicator found: '{tool_indicator}'")

    return metadata


def extract_metadata_xlrd(sheet) -> Dict[str, str]:
    """Extrae metadatos de las primeras filas (xlrd)."""
    metadata = {"web": "", "evaluators": "", "tool": "", "comment": "", "time": ""}

    # Buscar en primeras 10 filas
    for i in range(min(10, sheet.nrows)):
        try:
            row = sheet.row_values(i)
            if not row or not row[0]:
                continue

            key = str(row[0]).strip().lower()
            value = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            if "web site" in key:
                metadata["web"] = value
            elif "evaluator" in key:
                metadata["evaluators"] = value
            elif "tool used" in key:
                metadata["tool"] = value
            elif "comment" in key:
                metadata["comment"] = value
            elif "time spent" in key:
                metadata["time"] = value
        except:
            continue

    return metadata


def extract_metadata_before_table_xlrd(sheet, header_row: int) -> Dict[str, str]:
    """
    Extrae metadatos específicos de las filas ANTES de una tabla específica (xlrd).

    Args:
        sheet: La hoja de Excel
        header_row: Fila donde está el header de la tabla

    Returns:
        Diccionario con metadata incluyendo tool_indicator si se encuentra
    """
    metadata = {
        "web": "",
        "evaluators": "",
        "tool": "",
        "comment": "",
        "time": "",
        "tool_indicator": "",
    }

    # Buscar metadata en las 20 filas anteriores al header
    search_start = max(0, header_row - 20)

    for i in range(search_start, header_row):
        try:
            row = sheet.row_values(i)
            if not row or not row[0]:
                continue

            key = str(row[0]).strip().lower()
            value = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            if "web site" in key:
                metadata["web"] = value
            elif "evaluator" in key:
                metadata["evaluators"] = value
            elif "tool used" in key:
                metadata["tool"] = value
            elif "comment" in key:
                metadata["comment"] = value
            elif "time spent" in key:
                metadata["time"] = value
        except:
            continue

    # Buscar indicador de herramienta específico en la fila INMEDIATAMENTE anterior al header
    if header_row > 0:
        try:
            prev_row = sheet.row_values(header_row - 1)
            if len(prev_row) > 1 and prev_row[1]:
                tool_indicator = str(prev_row[1]).strip()
                # Verificar si contiene nombre de herramienta
                tool_lower = tool_indicator.lower()
                if any(
                    keyword in tool_lower
                    for keyword in ["gemini", "chatgpt", "chat gpt", "claude", "gpt"]
                ):
                    metadata["tool_indicator"] = tool_indicator
                    logger.debug(f"      Tool indicator found: '{tool_indicator}'")
        except:
            pass

    return metadata


def detect_multi_tool_columns_openpyxl(
    sheet, header_row_idx: int
) -> Tuple[bool, Optional[Dict[int, str]]]:
    """
    Detecta si hay múltiples columnas Value con nombres de herramientas (openpyxl).

    Returns:
        (is_multi_tool, tool_names_dict)
        - is_multi_tool: bool - True si hay múltiples herramientas
        - tool_names_dict: {col_idx: "tool_name"} - Mapeo de índice de columna a nombre de herramienta
    """
    # Obtener la fila de headers
    header_row = list(
        sheet.iter_rows(
            min_row=header_row_idx, max_row=header_row_idx, values_only=True
        )
    )[0]

    # Contar cuántas columnas tienen "Value" como header
    value_columns = []
    for idx, cell in enumerate(header_row):
        if cell and str(cell).strip().lower() == "value":
            value_columns.append(idx)

    # Si hay menos de 2 columnas "Value", no es multi-tool
    if len(value_columns) < 2:
        return False, None

    # Verificar la fila anterior para nombres de herramientas
    if header_row_idx <= 1:
        return False, None

    tool_row = list(
        sheet.iter_rows(
            min_row=header_row_idx - 1, max_row=header_row_idx - 1, values_only=True
        )
    )[0]

    # Buscar patrones "Tool X:" en las celdas correspondientes a las columnas Value
    tool_names = {}
    for col_idx in value_columns:
        if col_idx < len(tool_row) and tool_row[col_idx]:
            cell_text = str(tool_row[col_idx]).strip()
            # Verificar si contiene "tool" (case-insensitive)
            if "tool" in cell_text.lower():
                tool_names[col_idx] = cell_text

    # Si encontramos nombres de herramientas para al menos 2 columnas, es multi-tool
    if len(tool_names) >= 2:
        logger.debug(f"Multi-tool structure detected: {tool_names}")
        return True, tool_names

    return False, None


def find_data_start_row_openpyxl(sheet) -> Optional[int]:
    """Encuentra la fila donde comienzan los datos (openpyxl) - PRIMERA tabla."""
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_text = " ".join([str(cell).lower() if cell else "" for cell in row])
        if "elements to be identified" in row_text and "value" in row_text:
            return i
    return None


def find_all_data_start_rows_openpyxl(sheet) -> List[int]:
    """Encuentra TODAS las filas donde comienzan tablas (openpyxl)."""
    header_rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_text = " ".join([str(cell).lower() if cell else "" for cell in row])
        if "elements to be identified" in row_text and "value" in row_text:
            header_rows.append(i)
    return header_rows


def detect_multi_tool_columns_xlrd(
    sheet, header_row_idx: int
) -> Tuple[bool, Optional[Dict[int, str]]]:
    """
    Detecta si hay múltiples columnas Value con nombres de herramientas (xlrd).

    Returns:
        (is_multi_tool, tool_names_dict)
        - is_multi_tool: bool - True si hay múltiples herramientas
        - tool_names_dict: {col_idx: "tool_name"} - Mapeo de índice de columna a nombre de herramienta
    """
    try:
        # Obtener la fila de headers
        header_row = sheet.row_values(header_row_idx)

        # Contar cuántas columnas tienen "Value" como header
        value_columns = []
        for idx, cell in enumerate(header_row):
            if cell and str(cell).strip().lower() == "value":
                value_columns.append(idx)

        # Si hay menos de 2 columnas "Value", no es multi-tool
        if len(value_columns) < 2:
            return False, None

        # Verificar la fila anterior para nombres de herramientas
        if header_row_idx <= 0:
            return False, None

        tool_row = sheet.row_values(header_row_idx - 1)

        # Buscar patrones "Tool X:" en las celdas correspondientes a las columnas Value
        tool_names = {}
        for col_idx in value_columns:
            if col_idx < len(tool_row) and tool_row[col_idx]:
                cell_text = str(tool_row[col_idx]).strip()
                # Verificar si contiene "tool" (case-insensitive)
                if "tool" in cell_text.lower():
                    tool_names[col_idx] = cell_text

        # Si encontramos nombres de herramientas para al menos 2 columnas, es multi-tool
        if len(tool_names) >= 2:
            logger.debug(f"Multi-tool structure detected: {tool_names}")
            return True, tool_names

        return False, None
    except:
        return False, None


def find_data_start_row_xlrd(sheet) -> Optional[int]:
    """Encuentra la fila donde comienzan los datos (xlrd) - PRIMERA tabla."""
    for i in range(sheet.nrows):
        try:
            row = sheet.row_values(i)
            row_text = " ".join([str(cell).lower() if cell else "" for cell in row])
            if "elements to be identified" in row_text and "value" in row_text:
                return i
        except:
            continue
    return None


def find_all_data_start_rows_xlrd(sheet) -> List[int]:
    """Encuentra TODAS las filas donde comienzan tablas (xlrd)."""
    header_rows = []
    for i in range(sheet.nrows):
        try:
            row = sheet.row_values(i)
            row_text = " ".join([str(cell).lower() if cell else "" for cell in row])
            if "elements to be identified" in row_text and "value" in row_text:
                header_rows.append(i)
        except:
            continue
    return header_rows


def extract_data_rows_openpyxl(
    sheet, start_row: int, end_row: Optional[int] = None
) -> List[Dict[str, str]]:
    """
    Extrae filas de datos desde start_row (openpyxl).

    Args:
        sheet: Hoja de Excel
        start_row: Fila donde comienza el header de la tabla
        end_row: Fila opcional donde termina la tabla (inicio de la siguiente tabla)
    """
    data_rows = []
    current_category = "E2R Guidelines about Spelling"  # Default inicial

    # Detectar si hay estructura multi-herramienta
    is_multi_tool, tool_names = detect_multi_tool_columns_openpyxl(sheet, start_row)

    if is_multi_tool and tool_names:
        logger.info(f"    ⚙️  Multi-tool structure detected:")
        logger.info(f"       Tools: {list(tool_names.values())}")
        logger.info(f"       Each element will generate {len(tool_names)} rows")

    # Determinar el límite de filas a procesar
    max_row_to_process = end_row - 1 if end_row else sheet.max_row

    for row_idx, row in enumerate(
        sheet.iter_rows(
            min_row=start_row + 1, max_row=max_row_to_process, values_only=True
        ),
        start=start_row + 1,
    ):
        # Si llegamos al end_row, detener
        if end_row and row_idx >= end_row:
            break

        # Verificar si es fila vacía (fin de datos)
        if not any(cell for cell in row if cell):
            break

        # Columna B siempre es el elemento
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        # Si es una categoría E2R, actualizar current_category
        if "e2r guideline" in col_b.lower():
            current_category = col_b
            continue

        # Si no hay elemento, saltar
        if not col_b:
            continue

        if is_multi_tool and tool_names:
            # Estructura multi-herramienta: generar una fila por cada herramienta
            # tool_names es un dict {col_idx: "tool_name"}
            # Necesitamos extraer Value de cada columna correspondiente

            # Primero verificar si llegamos al final (hoja Values)
            first_value_col = min(tool_names.keys())
            first_value = (
                str(row[first_value_col]).strip()
                if len(row) > first_value_col and row[first_value_col]
                else ""
            )
            if "value" in col_b.lower() and "meaning" in first_value.lower():
                break

            # Encontrar la columna Details (después de las columnas Value)
            max_value_col = max(tool_names.keys())
            details_col_idx = max_value_col + 1
            comments_col_idx = max_value_col + 2

            col_details = (
                str(row[details_col_idx]).strip()
                if len(row) > details_col_idx and row[details_col_idx]
                else ""
            )
            col_comments = (
                str(row[comments_col_idx]).strip()
                if len(row) > comments_col_idx and row[comments_col_idx]
                else ""
            )

            # Generar una fila por cada herramienta
            for col_idx, tool_name in sorted(tool_names.items()):
                value = (
                    str(row[col_idx]).strip()
                    if len(row) > col_idx and row[col_idx]
                    else ""
                )

                data_rows.append(
                    {
                        "category": current_category,
                        "element": col_b,
                        "value": value,
                        "details": col_details,
                        "comments": col_comments,
                        "tool_specific": tool_name,  # Nombre específico de esta herramienta
                    }
                )
        else:
            # Estructura normal: solo una columna Value
            # Columnas: A(vacía), B(Element), C(Value), D(Details), E(Comments)
            col_c = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            col_d = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            col_e = str(row[4]).strip() if len(row) > 4 and row[4] else ""

            # Si encontramos "Value" + "Meaning" significa que es una hoja Values
            if "value" in col_b.lower() and "meaning" in col_c.lower():
                break

            data_rows.append(
                {
                    "category": current_category,
                    "element": col_b,
                    "value": col_c,
                    "details": col_d,
                    "comments": col_e,
                }
            )

    return data_rows


def extract_data_rows_xlrd(
    sheet, start_row: int, end_row: Optional[int] = None
) -> List[Dict[str, str]]:
    """
    Extrae filas de datos desde start_row (xlrd).

    Args:
        sheet: Hoja de Excel
        start_row: Fila donde comienza el header de la tabla
        end_row: Fila opcional donde termina la tabla (inicio de la siguiente tabla)
    """
    data_rows = []
    current_category = "E2R Guidelines about Spelling"  # Default inicial

    # Detectar si hay estructura multi-herramienta
    is_multi_tool, tool_names = detect_multi_tool_columns_xlrd(sheet, start_row)

    if is_multi_tool and tool_names:
        logger.info(f"    ⚙️  Multi-tool structure detected:")
        logger.info(f"       Tools: {list(tool_names.values())}")
        logger.info(f"       Each element will generate {len(tool_names)} rows")

    # Determinar el límite de filas a procesar
    max_row_to_process = end_row if end_row else sheet.nrows

    for i in range(start_row + 1, max_row_to_process):
        try:
            row = sheet.row_values(i)

            # Verificar si es fila vacía
            if not any(cell for cell in row if cell):
                break

            # Columna B siempre es el elemento
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            # Si es categoría E2R
            if "e2r guideline" in col_b.lower():
                current_category = col_b
                continue

            if not col_b:
                continue

            if is_multi_tool and tool_names:
                # Estructura multi-herramienta: generar una fila por cada herramienta

                # Primero verificar si llegamos al final (hoja Values)
                first_value_col = min(tool_names.keys())
                first_value = (
                    str(row[first_value_col]).strip()
                    if len(row) > first_value_col and row[first_value_col]
                    else ""
                )
                if "value" in col_b.lower() and "meaning" in first_value.lower():
                    break

                # Encontrar la columna Details (después de las columnas Value)
                max_value_col = max(tool_names.keys())
                details_col_idx = max_value_col + 1
                comments_col_idx = max_value_col + 2

                col_details = (
                    str(row[details_col_idx]).strip()
                    if len(row) > details_col_idx and row[details_col_idx]
                    else ""
                )
                col_comments = (
                    str(row[comments_col_idx]).strip()
                    if len(row) > comments_col_idx and row[comments_col_idx]
                    else ""
                )

                # Generar una fila por cada herramienta
                for col_idx, tool_name in sorted(tool_names.items()):
                    value = (
                        str(row[col_idx]).strip()
                        if len(row) > col_idx and row[col_idx]
                        else ""
                    )

                    data_rows.append(
                        {
                            "category": current_category,
                            "element": col_b,
                            "value": value,
                            "details": col_details,
                            "comments": col_comments,
                            "tool_specific": tool_name,  # Nombre específico de esta herramienta
                        }
                    )
            else:
                # Estructura normal: solo una columna Value
                col_c = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                col_d = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                col_e = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                if "value" in col_b.lower() and "meaning" in col_c.lower():
                    break

                data_rows.append(
                    {
                        "category": current_category,
                        "element": col_b,
                        "value": col_c,
                        "details": col_d,
                        "comments": col_e,
                    }
                )
        except:
            continue

    return data_rows


def process_sheet_openpyxl(
    sheet, sheet_name: str, file_path: Path, group_num: int
) -> List[Dict[str, str]]:
    """Procesa una hoja completa (openpyxl), incluyendo múltiples tablas si existen."""
    logger.info(f"  Processing sheet: {sheet_name}")

    # Encontrar TODAS las tablas en esta hoja
    all_table_rows = find_all_data_start_rows_openpyxl(sheet)

    if not all_table_rows:
        logger.warning(f"    No data headers found in {sheet_name}")
        return []

    logger.info(f"    Found {len(all_table_rows)} table(s) in {sheet_name}")

    result = []

    # Procesar cada tabla individualmente
    for idx, start_row in enumerate(all_table_rows, 1):
        logger.info(
            f"    Processing table {idx}/{len(all_table_rows)} starting at row {start_row}"
        )

        # Extraer metadatos específicos de esta tabla
        metadata = extract_metadata_before_table_openpyxl(sheet, start_row)

        # Determinar dónde termina esta tabla (antes de la siguiente o fin de hoja)
        if idx < len(all_table_rows):
            end_row = all_table_rows[idx] - 1  # Termina antes de la siguiente tabla
        else:
            end_row = sheet.max_row  # Última tabla va hasta el final

        # Extraer datos de esta tabla
        data_rows = extract_data_rows_openpyxl(sheet, start_row, end_row)

        # Combinar con metadatos
        for data in data_rows:
            # Prioridad: tool_indicator > tool_specific > tool
            if metadata.get("tool_indicator"):
                tool_used = metadata["tool_indicator"]
            elif data.get("tool_specific"):
                tool_used = data["tool_specific"]
            else:
                tool_used = metadata["tool"]

            result.append(
                {
                    "E2R Guidelines": data["category"],
                    "Elements to be identified": data["element"],
                    "Value": data["value"],
                    "Details": data["details"],
                    "Comments": data["comments"],
                    "Tool used": tool_used,
                    "Comment": metadata["comment"],
                    "Time Spent": metadata["time"],
                    "Group": str(group_num),
                    "Web": metadata["web"],
                }
            )

        logger.info(f"      Table {idx} extracted {len(data_rows)} rows")

    logger.info(f"    Total extracted {len(result)} rows from all tables")
    return result


def process_sheet_xlrd(
    sheet, sheet_name: str, file_path: Path, group_num: int
) -> List[Dict[str, str]]:
    """Procesa una hoja completa (xlrd), incluyendo múltiples tablas si existen."""
    logger.info(f"  Processing sheet: {sheet_name}")

    # Encontrar TODAS las tablas en esta hoja
    all_table_rows = find_all_data_start_rows_xlrd(sheet)

    if not all_table_rows:
        logger.warning(f"    No data headers found in {sheet_name}")
        return []

    logger.info(f"    Found {len(all_table_rows)} table(s) in {sheet_name}")

    result = []

    # Procesar cada tabla individualmente
    for idx, start_row in enumerate(all_table_rows, 1):
        logger.info(
            f"    Processing table {idx}/{len(all_table_rows)} starting at row {start_row}"
        )

        # Extraer metadatos específicos de esta tabla
        metadata = extract_metadata_before_table_xlrd(sheet, start_row)

        # Determinar dónde termina esta tabla (antes de la siguiente o fin de hoja)
        if idx < len(all_table_rows):
            end_row = all_table_rows[idx] - 1  # Termina antes de la siguiente tabla
        else:
            end_row = sheet.nrows  # Última tabla va hasta el final

        # Extraer datos de esta tabla
        data_rows = extract_data_rows_xlrd(sheet, start_row, end_row)

        # Combinar con metadatos
        for data in data_rows:
            # Prioridad: tool_indicator > tool_specific > tool
            if metadata.get("tool_indicator"):
                tool_used = metadata["tool_indicator"]
            elif data.get("tool_specific"):
                tool_used = data["tool_specific"]
            else:
                tool_used = metadata["tool"]

            result.append(
                {
                    "E2R Guidelines": data["category"],
                    "Elements to be identified": data["element"],
                    "Value": data["value"],
                    "Details": data["details"],
                    "Comments": data["comments"],
                    "Tool used": tool_used,
                    "Comment": metadata["comment"],
                    "Time Spent": metadata["time"],
                    "Group": str(group_num),
                    "Web": metadata["web"],
                }
            )

        logger.info(f"      Table {idx} extracted {len(data_rows)} rows")

    logger.info(f"    Total extracted {len(result)} rows from all tables")
    return result


def process_excel_file(file_path: Path, group_num: int) -> List[Dict[str, str]]:
    """Procesa todas las hojas de un archivo Excel."""
    all_rows = []

    try:
        # Primero intentar con openpyxl (funciona para .xlsx y algunos .xls mal nombrados)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                if is_part2_sheet(sheet, sheet_name):
                    rows = process_sheet_openpyxl(
                        sheet, sheet_name, file_path, group_num
                    )
                    all_rows.extend(rows)

            wb.close()
            return all_rows

        except Exception as e1:
            # Si falla openpyxl, intentar con xlrd (para .xls antiguos reales)
            if file_path.suffix.lower() == ".xls":
                try:
                    wb = xlrd.open_workbook(str(file_path))

                    for sheet_name in wb.sheet_names():
                        sheet = wb.sheet_by_name(sheet_name)

                        if is_part2_sheet(sheet, sheet_name):
                            rows = process_sheet_xlrd(
                                sheet, sheet_name, file_path, group_num
                            )
                            all_rows.extend(rows)

                    return all_rows
                except Exception as e2:
                    # Special case: xlrd says "Excel xlsx file; not supported"
                    # This means it's actually xlsx format with wrong .XLS extension
                    if "xlsx file; not supported" in str(e2).lower():
                        logger.info(
                            f"  Detected xlsx format with .XLS extension, renaming temporarily..."
                        )
                        import shutil

                        temp_path = file_path.with_suffix(".xlsx")
                        shutil.copy2(file_path, temp_path)
                        try:
                            wb = openpyxl.load_workbook(
                                temp_path, data_only=True, read_only=True
                            )

                            for sheet_name in wb.sheetnames:
                                sheet = wb[sheet_name]

                                if is_part2_sheet(sheet, sheet_name):
                                    rows = process_sheet_openpyxl(
                                        sheet, sheet_name, file_path, group_num
                                    )
                                    all_rows.extend(rows)

                            wb.close()
                            return all_rows
                        finally:
                            # Clean up temp file
                            if temp_path.exists():
                                temp_path.unlink()
                    else:
                        logger.error(f"  Error processing {file_path.name}: {e2}")
            else:
                logger.error(f"  Error processing {file_path.name}: {e1}")

    except Exception as e:
        logger.error(f"  Error processing {file_path.name}: {e}")

    return all_rows


def find_all_part2_data() -> Tuple[List[Dict[str, str]], Dict[str, int], List[str]]:
    """
    Busca todos los archivos y hojas con datos Part II.

    Returns:
        (all_rows, group_mapping, groups_without_data)
    """
    base_dir = Path("/home/javi/practices/practices")
    student_dirs = sorted([d for d in base_dir.glob("*_assignsubmission_file")])

    # Crear mapeo de grupos
    group_mapping = {d.name: idx + 1 for idx, d in enumerate(student_dirs)}

    all_rows = []
    groups_without_data = []

    logger.info(f"Found {len(student_dirs)} student directories")
    logger.info("=" * 80)

    for student_dir in student_dirs:
        student_name = student_dir.name.split("_")[0]
        group_num = group_mapping[student_dir.name]

        logger.info(f"\nGroup {group_num}: {student_name}")

        # Buscar todos los archivos Excel (case-insensitive)
        xlsx_files = list(student_dir.glob("*.xlsx")) + list(student_dir.glob("*.XLSX"))
        xls_files = list(student_dir.glob("*.xls")) + list(student_dir.glob("*.XLS"))
        all_excel_files = xlsx_files + xls_files

        if not all_excel_files:
            logger.warning(f"  No Excel files found")
            groups_without_data.append(f"{group_num}. {student_name}")
            continue

        student_rows = []
        for excel_file in all_excel_files:
            logger.info(f"  File: {excel_file.name}")
            rows = process_excel_file(excel_file, group_num)
            student_rows.extend(rows)

        if student_rows:
            all_rows.extend(student_rows)
            logger.info(f"  ✓ Total rows for this group: {len(student_rows)}")
        else:
            logger.warning(f"  ✗ No Part II data found")
            groups_without_data.append(f"{group_num}. {student_name}")

    return all_rows, group_mapping, groups_without_data


def write_csv(all_rows: List[Dict[str, str]], output_path: Path):
    """Escribe el CSV consolidado."""
    if not all_rows:
        logger.error("No data to write!")
        return

    headers = [
        "E2R Guidelines",
        "Elements to be identified",
        "Value",
        "Details",
        "Comments",
        "Tool used",
        "Comment",
        "Time Spent",
        "Group",
        "Web",
    ]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(all_rows)

    logger.info(f"\n✅ CSV written to: {output_path}")
    logger.info(f"   Total rows: {len(all_rows)}")


def write_report(
    all_rows: List[Dict[str, str]],
    group_mapping: Dict[str, int],
    groups_without_data: List[str],
    report_path: Path,
):
    """Genera reporte de consolidación."""
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("=" * 80 + "\n")
        f.write("REPORTE DE CONSOLIDACIÓN - PRÁCTICA II\n")
        f.write("=" * 80 + "\n\n")

        f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        f.write("ESTADÍSTICAS:\n")
        f.write(f"  • Total de grupos: {len(group_mapping)}\n")
        f.write(
            f"  • Grupos con datos: {len(group_mapping) - len(groups_without_data)}\n"
        )
        f.write(f"  • Grupos sin datos: {len(groups_without_data)}\n")
        f.write(f"  • Total de filas CSV: {len(all_rows)}\n\n")

        if groups_without_data:
            f.write("⚠️  GRUPOS SIN PRÁCTICA II:\n")
            for group in groups_without_data:
                f.write(f"  {group}\n")
            f.write("\n")

        # Distribución por grupo
        f.write("\nDISTRIBUCIÓN DE FILAS POR GRUPO:\n")
        group_counts = {}
        for row in all_rows:
            group = row["Group"]
            group_counts[group] = group_counts.get(group, 0) + 1

        for group_num in sorted(group_counts.keys(), key=int):
            count = group_counts[group_num]
            f.write(f"  Grupo {group_num}: {count} filas\n")

        f.write("\n" + "=" * 80 + "\n")

    logger.info(f"📄 Report written to: {report_path}")


def main():
    """Función principal."""
    logger.info("Starting Part II consolidation...")
    logger.info("=" * 80)

    # Buscar todos los datos
    all_rows, group_mapping, groups_without_data = find_all_part2_data()

    logger.info("\n" + "=" * 80)
    logger.info("CONSOLIDATION COMPLETE")
    logger.info("=" * 80)

    if not all_rows:
        logger.error("❌ No data found to consolidate!")
        return

    # Escribir CSV
    csv_path = ROOT_DIR / "data/part2_consolidated.csv"
    write_csv(all_rows, csv_path)

    # Escribir reporte
    report_path = ROOT_DIR / "logs/part2_consolidation_report.txt"
    write_report(all_rows, group_mapping, groups_without_data, report_path)

    logger.info("\n✅ All done!")
    logger.info(f"   CSV: {csv_path}")
    logger.info(f"   Report: {report_path}")
    logger.info(f"   Log: logs/part2_consolidation.log")


if __name__ == "__main__":
    main()
